import inspect
import logging
from openpyxl import load_workbook

# create logger
class Utils:
    def StartLog(self,logLevel = logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name) #__name__ is the name of the logger that we create
        logger.setLevel(logLevel)
        # create file handler
        fh = logging.FileHandler('../Tests_log.log')  # file handler
        # create formatter
        formatter = logging.Formatter('%(asctime)s %(levelname)s [%(name)s] %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to handler
        fh.setFormatter(formatter)
        # add handlers to logger
        logger.addHandler(fh)
        return logger

    def readExcelFile(fname, worksheet=None):
        datalist = []  # list of data from the file
        wb = load_workbook(filename=fname)
        if worksheet is None:
            worksheet = wb.active
        else:
            worksheet = wb[worksheet]
        totalRows = worksheet.max_row  # quantity of rows in a sheet
        totalCols = worksheet.max_column  # quantity of cols in a sheet
        for y in range(2, totalRows + 1):
            row = []
            for x in range(1, totalCols + 1):
                row.append(worksheet.cell(y, x).value)
            datalist.append(row)
        return datalist
